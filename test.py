import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# Fonction pour charger les données à partir d'un fichier CSV
def load_data(file):
    file_path = f"Datas/{file}.csv"
    if os.path.exists(file_path):
        return pd.read_csv(file_path)
    else:
        return pd.DataFrame()

# Fonction pour enregistrer les données dans un fichier CSV
def save_data(file, data):
    file_path = f"Datas/{file}.csv"
    data.to_csv(file_path, index=False)

def merge_convention_logs(convention_file_path, logs_file_path):
    # Charger les données des fichiers convention et logs
    convention_df = pd.read_csv(convention_file_path)
    logs_df = pd.read_csv(logs_file_path)

    # Fusionner les données en utilisant une jointure gauche sur la colonne 'Libellé'
    merged_df = pd.merge(convention_df, logs_df, on='Libellé', how='left') 
    
    return merged_df # final_df
 
def export_to_excel(data_frame, output_file_path):
    # Obtenir la date d'aujourd'hui
    date_aujourdhui = datetime.today()

    # Formater la date au format DD/MM/YY
    date_formattee = date_aujourdhui.strftime("%d/%m/%y")   

    print(data_frame.columns)
    # Créer un nouveau classeur Excel
    wb = Workbook()

    #Créer une page pour le bilan
    bilan_ws=wb.create_sheet(title="BILAN")
    
    #En-tête du BILAN 
    bilan_ws['A3'] = 'BILAN CONVENTIONS CREST / ÉCOLE POLYTECHNIQUE'
    bilan_ws['A4'] = " ".join(['MAJ',date_formattee])
    bilan_ws['A7'] = 'CONVENTIONS'
    bilan_ws['B7'] = 'MONTANTS ALLOUÉS'
    bilan_ws['F7'] = 'DÉPENSES'
    bilan_ws['J7'] = 'DISPONIBLES' 
    bilan_ws['B8'] = 'Fonctionnement'
    bilan_ws['C8'] = 'Investissement'
    bilan_ws['D8'] = 'Personnel'
    bilan_ws['E8'] = 'Total' 
    bilan_ws['F8'] = 'Fonctionnement'
    bilan_ws['G8'] = 'Investissement'
    bilan_ws['H8'] = 'Personnel'
    bilan_ws['I8'] = 'Total'
    bilan_ws['J8'] = 'Fonctionnement'
    bilan_ws['K8'] = 'Investissement'
    bilan_ws['L8'] = 'Personnel'
    bilan_ws['M8'] = 'Total'

    # Initialiser le compteur de ligne pour le bilan
    ligne_bilan = 8
    # Parcourir chaque convention et écrire ses données dans une feuille Excel distincte
    for convention in data_frame['nPEP_x'].unique():
        # Sélectionner les données de la convention actuelle
        convention_data = data_frame[data_frame['nPEP_x'] == convention]  

        # Créer une nouvelle feuille avec le nom de la convention
        ws = wb.create_sheet(title=convention)

        # Faire un en-tête 
        ws['A1']='Date MAJ'
        ws['B1']=date_formattee

        #Tableau FICHE IDENTITE CONVENTION
        ws['B5']='FICHE IDENTITE CONVENTION'
        ws['B6']='Libellé :'
        ws['B8']='Dates :'
        ws['B9']='Bénéficiaire :'
        ws['B10']='Montant alloué :'
        ws['B11']='Fonctionnement'
        ws['B12']='Investissement'
        ws['B13']='Personnel'
        ws['E6']='n° PEP :'
        ws['E8']='Partenaire :'
        ws['E9']='Disponible :' 
        ws['E10']='Fonctionnement'
        ws['E11']='Investissement'
        ws['E12']='Personnel'
        #valeur
        ws['C6']=convention_data.iloc[0]['Libellé']
        ws['C8']=" - ".join([convention_data.iloc[0]['Dates_d'], convention_data.iloc[0]['Dates_f']]) #+'-'+convention_data.iloc[0]['Dates_f']
        ws['C9']=convention_data.iloc[0]['Bénéficiaire_x']
        ws['C10']=convention_data.iloc[0]['Montant alloué'] 
        ws['C11']=convention_data.iloc[0]['M_Fonctionnement']  
        ws['C12']=convention_data.iloc[0]['M_Investissement'] 
        ws['C13']=convention_data.iloc[0]['M_Personnel']
        ws['F6']=convention_data.iloc[0]['nPEP_x']
        ws['F8']=convention_data.iloc[0]['Partenaire'] 
        # Calculer la somme totale des valeurs dans la colonne 'Montant de la dépense (HTR)'
        montant_total = convention_data['Montant de la dépense (HTR)'].sum()
        # Calculer la somme des valeurs dans la colonne 'Montant de la dépense (HTR)' lorsque 'Nature de la dépense' est égale à 'Fonctionnement'
        montant_fonctionnement = convention_data.loc[convention_data['Nature de la dépense'] == 'Fonctionnement', 'Montant de la dépense (HTR)'].sum()
        # Calculer la somme des valeurs dans la colonne 'Montant de la dépense (HTR)' lorsque 'Nature de la dépense' est égale à 'Investissement'
        montant_investissement = convention_data.loc[convention_data['Nature de la dépense'] == 'Investissement', 'Montant de la dépense (HTR)'].sum()
        # Calculer la somme des valeurs dans la colonne 'Montant de la dépense (HTR)' lorsque 'Nature de la dépense' est égale à 'Personnel'
        montant_personnel = convention_data.loc[convention_data['Nature de la dépense'] == 'Personnel', 'Montant de la dépense (HTR)'].sum()

        # Placer les sommes dans les cellules appropriées
        ws['F9'] = convention_data.iloc[0]['Montant alloué'] - montant_total
        ws['F10'] = convention_data.iloc[0]['M_Fonctionnement'] - montant_fonctionnement
        ws['F11']=convention_data.iloc[0]['M_Investissement'] - montant_investissement
        ws['F12']=convention_data.iloc[0]['M_Personnel'] - montant_personnel 
        
        # Écrire les données dans la feuille Excel à partir de la ligne 19
        for r_idx, row in enumerate(dataframe_to_rows(convention_data[['Nature de la dépense', "Année d'imputation", 'Type de dépense', 'Bénéficiaire_y', 'Libellé.1', 'Numéro OM', 'Numéro EJ', 'Numéro Ligne', 'Montant de la dépense (HTR)', 'Solde']], index=False, header=True), 19):
            # Renommer les colonnes
            row = ['Bénéficiaire' if col == 'Bénéficiaire.y' else 'Libellé' if col == 'Libellé.1' else col for col in row]
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        #Compléter la page de BILAN 
        bilan_ws[f'A{ligne_bilan+1}']=convention_data.iloc[0]['Libellé']
        bilan_ws[f'B{ligne_bilan+1}']=convention_data.iloc[0]['M_Fonctionnement']
        bilan_ws[f'C{ligne_bilan+1}']=convention_data.iloc[0]['M_Investissement']
        bilan_ws[f'D{ligne_bilan+1}']=convention_data.iloc[0]['M_Personnel']
        bilan_ws[f'E{ligne_bilan+1}']=convention_data.iloc[0]['Montant alloué']
        bilan_ws[f'F{ligne_bilan+1}']=montant_fonctionnement
        bilan_ws[f'G{ligne_bilan+1}']=montant_investissement
        bilan_ws[f'H{ligne_bilan+1}']=montant_personnel 
        bilan_ws[f'I{ligne_bilan+1}']=montant_total
        bilan_ws[f'J{ligne_bilan+1}']=convention_data.iloc[0]['M_Fonctionnement'] - montant_fonctionnement
        bilan_ws[f'K{ligne_bilan+1}']=convention_data.iloc[0]['M_Investissement'] - montant_investissement
        bilan_ws[f'L{ligne_bilan+1}']=convention_data.iloc[0]['M_Personnel'] - montant_personnel 
        bilan_ws[f'M{ligne_bilan+1}']=convention_data.iloc[0]['Montant alloué']- montant_total

        #La prochaine convention sera sur la ligne suivante 
        ligne_bilan = ligne_bilan +1

    # Supprimer la feuille par défaut "Sheet"
    wb.remove(wb['Sheet'])

    # Enregistrer le classeur Excel
    wb.save(output_file_path)

    # Afficher un message de succès
    st.success(f"Données exportées avec succès vers {output_file_path}")

def convert_excel_to_convention_logs(excel_file_path,output_file_path):
    # Lire le fichier Excel
    xls = pd.ExcelFile(excel_file_path)

    # Lire chaque feuille de l'Excel et les convertir en DataFrames
    first_row_convention =["Libellé","Dates_d","Dates_f","nPEP","Bénéficiaire","Partenaire","Montant alloué","M_Fonctionnement","M_Investissement","M_Personnel"]
    first_row_log =["Date_ajout","Libellé","nPEP","Nature de la dépense","Année d'imputation","Type de dépense","Bénéficiaire","Libellé","Numéro OM","Numéro EJ",
                    "Numéro Ligne","Montant de la dépense (HTR)","Solde"]
    
    # Transposer les listes pour créer les DataFrames
    convention_df = pd.DataFrame([], columns=first_row_convention)
    logs_df = pd.DataFrame([], columns=first_row_log)

    for sheet_name in xls.sheet_names:
        if sheet_name != "BILAN" and sheet_name != "DATA":
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

            date=df.iloc[6,2] 
            date_parts = date.split(' - ') 

            #Remplir la ligne suivante avec toutes les informations            
            convention_row=[df.iloc[4,2],date_parts[0].strip(),date_parts[1].strip(),df.iloc[4,5],df.iloc[7,2],df.iloc[6,5],df.iloc[8,2],df.iloc[9,2],df.iloc[10,2],df.iloc[11,2]]
            conv_row_df=pd.DataFrame([convention_row],columns=first_row_convention) 
            i=18
            logs_row_df=pd.DataFrame([], columns=first_row_log) 
            while not pd.isnull(df.iloc[i, 0]) : 
                logs_row=["",df.iloc[4,2],df.iloc[4,5],df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],str(df.iloc[i,4]).replace('\n', ' '),df.iloc[i,5],df.iloc[i,6],
                    df.iloc[i,7],df.iloc[i,8],df.iloc[i,9]] 
                new_logs_row_df=pd.DataFrame([logs_row], columns=first_row_log)
                logs_row_df=pd.concat([logs_row_df,new_logs_row_df], ignore_index=True)
                i=i+1

            convention_df=pd.concat([convention_df, conv_row_df], ignore_index=True) 
            logs_df=pd.concat([logs_df, logs_row_df], ignore_index=True)  

    # Enregistrer les DataFrames en fichiers CSV
    convention_df.to_csv(output_file_path[0], index=False)
    logs_df.to_csv(output_file_path[1], index=False)

    return "Fichiers CSV pour conventions et logs créés avec succès !"

########################################        
### Menu principal et fonctionnalité ###
########################################

def main():
    #Authentification
    if not st.session_state.get("logged_in"):
        login_page()
        return

    st.sidebar.title("Navigation") 
    
    # Afficher les boutons pour chaque page
    if st.session_state.access_level == "Admin":
        page = st.sidebar.radio("Aller à :", ("Accueil", "Créer une convention", "Modifier une convention", "Enregistrer une opération", "Modifier une opération sur une convention", "Fusionner dans un excel", "Convertir en log"))
    elif st.session_state.access_level == "Gueux":
        page = st.sidebar.radio("Aller à :", ( "Enregistrer une opération", "Modifier une opération sur une convention"))

    # Afficher la page sélectionnée
    if page == "Accueil":
        accueil_page()
    elif page == "Créer une convention":
        creer_convention_page()
    elif page == "Modifier une convention":
        modifier_convention_page()
    elif page == "Enregistrer une opération":
        enregistrer_operation_page()
    elif page == "Modifier une opération sur une convention":
        modifier_logs_page()
    elif page == "Fusionner dans un excel":
        merge_convention_logs_page()
    elif page == "Convertir en log":
        Convert_excel_to_convention_logs_page()

#####################################        
### Contenu des différentes pages ###
#####################################

#Authentique 
def login_page():
    st.title("Authentification")
    username = st.text_input("Nom d'utilisateur")
    password = st.text_input("Mot de passe", type="password") 
    # Charger les données des utilisateurs à partir du fichier CSV
    users_info = pd.read_csv("Datas/acces.csv")
    
    # Afficher les informations des utilisateurs (pour le débogage)
    st.write(users_info)
    
    if st.button("Se Connecter"):
        # Vérifier si l'utilisateur et le mot de passe existent dans les données chargées
        if ((users_info['Id'] == username) & (users_info['password'] == password)).any():
            st.success("Connexion réussie !")
            st.session_state.user_id = username
            st.session_state.access_level = users_info.loc[users_info['Id'] == username, 'level_access'].iloc[0]
            st.session_state.logged_in = True  # Marquer l'utilisateur comme connecté
            st.rerun()
        else:
            st.error("Identifiants invalides.")


# Fonction pour afficher la page "Accueil"
def accueil_page():
    st.title("Page d'Accueil")
    convention_df = load_data("convention")
    st.subheader("Contenu du fichier convention :")
    st.write(convention_df)

# Fonction pour afficher la page "Créer une convention"
def creer_convention_page():
    st.title("Créer une convention")
    st.write("Veuillez remplir les informations suivantes :")

    # Obtenir les noms de colonnes du DataFrame
    convention_df = load_data("convention")
    column_names = convention_df.columns.tolist()
    user_inputs = {}

    # Afficher les champs de saisie pour chaque colonne
    for column in column_names:
        user_input = st.text_input(f"Saisir {column}", "")
        user_inputs[column] = user_input

    # Bouton pour enregistrer la convention
    if st.button("Enregistrer"):
        # Vérifier que tous les champs sont remplis
        if all(user_inputs.values()):
            # Vérifier si le couple "numéro PEP" et "Libellé" est déjà présent dans les conventions enregistrées
            existing_conventions = convention_df[(convention_df['nPEP'] == user_inputs['nPEP']) & (convention_df['Libellé'] == user_inputs['Libellé'])]
            if not existing_conventions.empty:
                st.warning("Une convention avec ce numéro PEP et ce libellé existe déjà. Veuillez enregistrer une convention unique.")
            else:
                # Créer un DataFrame avec les nouvelles données
                new_data = pd.DataFrame([user_inputs])
                save_data("convention", pd.concat([convention_df, new_data], ignore_index=True))
                st.success("Convention enregistrée avec succès !")
        else:
            st.warning("Veuillez remplir tous les champs avant d'enregistrer la convention.")

# Fonction pour afficher la page "Enregistrer une opération"
def enregistrer_operation_page():
    st.title("Enregistrer une opération sur une convention")
    st.write("Veuillez remplir les informations suivantes :")

    # Charger les données des conventions
    convention_df = load_data("convention")
    
    # Créer une liste de conventions avec les numéros PEP pour le menu déroulant
    convention_names = [f"{libelle} ({nPEP})" for libelle, nPEP in zip(convention_df['Libellé'], convention_df['nPEP'])]

    # Menu déroulant pour choisir la convention
    selected_convention = st.selectbox("Choisir une convention :", convention_names, index=0, format_func=lambda x: x.title(), key="convention_selectbox", help="Commencez à taper pour rechercher")

    # Extraire le numéro PEP et la convention de la convention sélectionnée
    selected_convention1 = selected_convention.split(' (')[0]
    selected_pep_number = selected_convention.split(' (')[1].split(')')[0]

    # Obtenir les noms de colonnes du DataFrame des opérations
    logs_df = load_data("logs")
    column_names = logs_df.columns.tolist()
    user_inputs = {}

    # Afficher les champs de saisie pour chaque colonne
    for column in column_names:
        if column == 'Date':  # Remplir automatiquement la date
            user_input = st.text_input("Date", value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), key="date_input", disabled=True)
            user_inputs['Date'] = user_input
        elif column == 'nPEP':  # Afficher le numéro PEP de la convention sélectionnée
            st.text_input("nPEP", value=selected_pep_number, key="pep_number_input", disabled=True)
            user_inputs['nPEP'] = selected_pep_number
        else:
            user_input = st.text_input(f"Saisir {column}", "")
            user_inputs[column] = user_input

    # Bouton pour enregistrer l'opération
    if st.button("Enregistrer"):
        # Vérifier que tous les champs sont remplis
        if all(user_inputs.values()):
            # Créer un DataFrame avec les nouvelles données
            new_data = pd.DataFrame([user_inputs])
            save_data("logs", pd.concat([logs_df, new_data], ignore_index=True))
            st.success("Opération sur la convention enregistrée avec succès !")
        else:
            st.warning("Veuillez remplir tous les champs avant d'enregistrer l'opération.")

# Fonction pour afficher la page "Modifier une convention"
def modifier_convention_page():
    st.title("Modifier une convention")

    # Charger les données des conventions
    convention_df = load_data("convention")

    # Afficher les conventions actuelles dans un tableau
    st.subheader("Conventions actuelles :")
    st.write(convention_df)

    # Sélectionner une ligne à modifier
    selected_index = st.number_input("Entrez l'index de la ligne à modifier :", min_value=0, max_value=len(convention_df)-1, value=0, step=1)

    # Définir une variable de session pour suivre l'état du bouton "Afficher la ligne sélectionnée"
    if 'Btn_affiche_clicked' not in st.session_state:
        st.session_state.Btn_affiche_clicked = False

    # Fonction pour gérer le clic sur le bouton "Afficher la ligne sélectionnée"
    def click_button():
        st.session_state.Btn_affiche_clicked = True

    # Afficher les boutons dans une colonne Bootstrap
    col1, col2 = st.columns(2)

    # Bouton "Afficher la ligne sélectionnée"
    with col1:
        if st.button("Modifier la ligne sélectionnée", on_click=click_button):
            st.session_state.Btn_affiche_clicked = True

    # Bouton "Supprimer la ligne sélectionnée"
    with col2:
        if st.button("Supprimer la ligne sélectionnée", key="supprimer_ligne"):
            # Supprimer la ligne sélectionnée du DataFrame
            convention_df.drop(selected_index, inplace=True)
            # Réinitialiser les index du DataFrame
            convention_df.reset_index(drop=True, inplace=True)
            # Enregistrer les modifications dans le fichier convention.csv
            save_data("convention", convention_df)
            st.success("Ligne supprimée avec succès !")
            st.rerun()

    # Vérifier si le bouton a été cliqué
    if st.session_state.Btn_affiche_clicked:
        # Afficher les détails de la ligne sélectionnée
        selected_row = convention_df.iloc[selected_index]
        st.write("Détails de la ligne sélectionnée :")
        st.write(selected_row)

        # Afficher les champs de saisie pour modifier la ligne
        st.subheader("Modifier les informations :")
        new_values = {}
        for column in convention_df.columns:
            new_value = st.text_input(f"Modifier {column} :", value=selected_row[column], key=f"{column}_{selected_index}")
            new_values[column] = new_value

        # Bouton pour enregistrer les modifications
        if st.button("Enregistrer les modifications", key="enregistrer_modifs"):
            # Mettre à jour la ligne sélectionnée dans le DataFrame
            convention_df.loc[selected_index] = pd.Series(new_values)

            # Enregistrer les modifications dans le fichier convention.csv
            save_data("convention", convention_df)

            st.success("Modification enregistrée avec succès !")
            # Rafraîchir la page
            st.experimental_rerun()


# Fonction pour afficher la page "Modifier une opération sur les conventions"
def modifier_logs_page():
    st.title("Modifier une opération sur les conventions")

    # Charger les données des opérations sur les conventions
    logs_df = load_data("logs")

    # Afficher les Opérations sur les conventions dans un tableau
    st.subheader("Opérations sur les conventions :")
    st.write(logs_df)

    # Sélectionner une ligne à modifier
    selected_index = st.number_input("Entrez l'index de la ligne à modifier :", min_value=0, max_value=len(logs_df)-1, value=0, step=1)

    
    if 'Btn_affiche_clicked' not in st.session_state:
        st.session_state.Btn_affiche_clicked = False
    
    def click_button():
        st.session_state.Btn_affiche_clicked = True
        
    # Afficher les boutons dans une colonne Bootstrap
    col1, col2 = st.columns(2)

    # Bouton "Afficher la ligne sélectionnée"
    with col1:
        if st.button("Modifier la ligne sélectionnée", on_click=click_button):
            st.session_state.Btn_affiche_clicked = True

    # Bouton "Supprimer la ligne sélectionnée"
    with col2:
        if st.button("Supprimer la ligne sélectionnée", key="supprimer_ligne"):
            # Supprimer la ligne sélectionnée du DataFrame
            logs_df.drop(selected_index, inplace=True)
            # Réinitialiser les index du DataFrame
            logs_df.reset_index(drop=True, inplace=True)
            # Enregistrer les modifications dans le fichier logs.csv
            save_data("logs", logs_df)
            st.success("Opération supprimée avec succès !")

    if st.session_state.Btn_affiche_clicked:
        
        # Afficher les détails de la ligne sélectionnée
        selected_row = logs_df.iloc[selected_index]
        st.write("Détails de la ligne sélectionnée :")
        st.write(selected_row)

        # Afficher les champs de saisie pour modifier la ligne
        st.subheader("Modifier les informations :")
        new_values = {}
        for column in logs_df.columns:
            new_value = st.text_input(f"Modifier {column} :", value=selected_row[column], key=f"{column}_{selected_index}")
            new_values[column] = new_value

        # Bouton pour enregistrer les modifications
        if st.button("Enregistrer les modifications", key="enregistrer_modifs"):
            # Mettre à jour la ligne sélectionnée dans le DataFrame
            logs_df.loc[selected_index] = pd.Series(new_values)

            # Enregistrer les modifications dans le fichier logs.csv
            save_data("logs", logs_df)

            st.success("Modification enregistrée avec succès !")
            #Refresh the page
            st.rerun() 

# Fonction pour afficher la page "Modifier une opération sur les conventions"
def merge_convention_logs_page():
    st.title("Fusion des données Convention et Logs")

    # Afficher les champs de téléchargement des fichiers
    convention_file = load_data("convention")
    logs_file = load_data("logs")

    if convention_file is not None and logs_file is not None:
        # Enregistrer les fichiers téléchargés temporairement
        convention_temp_path = convention_file
        logs_temp_path = logs_file 

        # Fusionner les fichiers et calculer le bilan des dépenses
        merged_data = merge_convention_logs("Datas/convention.csv", "Datas/logs.csv")

        # Sélectionner le chemin de sortie pour le fichier Excel
        output_file_path = st.text_input("Nom du fichier Excel de sortie :", "conventions_logs_output.xlsx")

        # Bouton pour exporter les données fusionnées vers un fichier Excel
        if st.button("Exporter vers Excel"):
            export_to_excel(merged_data, "Datas/"+output_file_path)
            st.success(f"Données exportées avec succès vers {output_file_path}")

        # Afficher les données fusionnées dans Streamlit
        st.subheader("Données fusionnées Convention et Logs :")
        st.write(merged_data) 


# Fonction pour Convertir un excel en fichier convention et log
def Convert_excel_to_convention_logs_page(): 
    st.title("Fusion des données Convention et Logs")

    # Afficher les champs de téléchargement des fichiers
    convention_file = st.file_uploader("Sélectionner le fichier convention.excel", type=['xlsx'])

    # Sélectionner le chemin de sortie pour le fichier Excel
    output_conv_file_path = st.text_input("Nom du fichier des conventions csv de sortie :", "conventions1.csv")
    output_log_file_path = st.text_input("Nom du fichier des opérations csv de sortie :", "logs1.csv")

    if st.button("Exporter vers Excel"):        
        first_page_df = pd.read_excel(convention_file, sheet_name='SUBVENTION')
        st.write(first_page_df) 
        convert_excel_to_convention_logs(convention_file,["Datas/"+output_conv_file_path,"Datas/"+output_log_file_path])



# Appeler la fonction principale
if __name__ == "__main__":
    main()
